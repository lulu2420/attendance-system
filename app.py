“””
نظام الحضور والغياب — واجهة ويب
Flask + SQLite + TTLock API
“””
from flask import Flask, render_template, request, jsonify, send_file
from apscheduler.schedulers.background import BackgroundScheduler
import sqlite3, hashlib, time, requests, smtplib, os, io
from datetime import datetime, date, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(**name**)
DB = “attendance.db”

# ============================================================

# إعدادات — عدّل هذه فقط

# ============================================================

TTLOCK_CONFIG = {
“client_id”:     os.environ.get(“TTLOCK_CLIENT_ID”,     “YOUR_CLIENT_ID”),
“client_secret”: os.environ.get(“TTLOCK_CLIENT_SECRET”, “YOUR_CLIENT_SECRET”),
“username”:      os.environ.get(“TTLOCK_USERNAME”,      “YOUR_EMAIL”),
“password”:      os.environ.get(“TTLOCK_PASSWORD”,      “YOUR_PASSWORD”),
}
EMAIL_CONFIG = {
“sender_email”:    os.environ.get(“EMAIL_SENDER”,   “your@gmail.com”),
“sender_password”: os.environ.get(“EMAIL_PASSWORD”, “your_app_password”),
“smtp_server”:     “smtp.gmail.com”,
“smtp_port”:       587,
}

# ============================================================

# قاعدة البيانات

# ============================================================

def get_db():
conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
return conn

def init_db():
with get_db() as db:
db.executescript(”””
CREATE TABLE IF NOT EXISTS employees (
id INTEGER PRIMARY KEY AUTOINCREMENT,
name_en TEXT UNIQUE NOT NULL,
name_ar TEXT NOT NULL,
email TEXT,
salary REAL DEFAULT 0,
type TEXT DEFAULT ‘fixed’,
shift_start TEXT DEFAULT ‘09:00’,
shift_end TEXT DEFAULT ‘17:00’,
weekly_hours INTEGER DEFAULT 40,
housing_allowance REAL DEFAULT 0,
transport_allowance REAL DEFAULT 0,
commission REAL DEFAULT 0,
other_deductions REAL DEFAULT 0,
lock_id TEXT,
active INTEGER DEFAULT 1
);
CREATE TABLE IF NOT EXISTS attendance (
id INTEGER PRIMARY KEY AUTOINCREMENT,
emp_name_en TEXT,
date TEXT,
check_in TEXT,
check_out TEXT,
delay_minutes REAL DEFAULT 0,
early_minutes REAL DEFAULT 0,
violation_in TEXT,
violation_out TEXT,
deduction_sar REAL DEFAULT 0,
warning INTEGER DEFAULT 0,
created_at TEXT DEFAULT (datetime(‘now’))
);
CREATE TABLE IF NOT EXISTS payroll (
id INTEGER PRIMARY KEY AUTOINCREMENT,
emp_name_en TEXT,
month TEXT,
basic REAL,
housing REAL,
transport REAL,
commission REAL,
gross REAL,
att_deduction REAL,
other_deductions REAL,
net REAL,
created_at TEXT DEFAULT (datetime(‘now’))
);
“””)

# ============================================================

# TTLock

# ============================================================

_token_cache = {“token”: None, “expiry”: 0}

def get_ttlock_token():
if _token_cache[“token”] and time.time() < _token_cache[“expiry”]:
return _token_cache[“token”]
cfg = TTLOCK_CONFIG
pw  = hashlib.md5(cfg[“password”].encode()).hexdigest()
r   = requests.post(“https://euapi.ttlock.com/oauth2/token”, data={
“clientId”: cfg[“client_id”], “clientSecret”: cfg[“client_secret”],
“username”: cfg[“username”],  “password”: pw, “grant_type”: “password”
}, timeout=10).json()
if “access_token” in r:
_token_cache[“token”]  = r[“access_token”]
_token_cache[“expiry”] = time.time() + r.get(“expires_in”, 7200) - 60
return _token_cache[“token”]
raise Exception(f”TTLock auth failed: {r}”)

def fetch_ttlock_records(start: date, end: date):
token    = get_ttlock_token()
start_ms = int(datetime.combine(start, datetime.min.time()).timestamp() * 1000)
end_ms   = int(datetime.combine(end,   datetime.max.time()).timestamp() * 1000)
r = requests.get(“https://euapi.ttlock.com/v3/lockRecord/list”, params={
“clientId”: TTLOCK_CONFIG[“client_id”], “accessToken”: token,
“startDate”: start_ms, “endDate”: end_ms,
“pageNo”: 1, “pageSize”: 200, “date”: int(time.time()*1000)
}, timeout=10).json()
return r.get(“list”, [])

def parse_daily(records, employees):
result = {e[“name_en”]: {} for e in employees}
for rec in records:
username = rec.get(“username”, “”)
ts = rec.get(“lockDate”, rec.get(“successDate”, 0))
if not ts: continue
dt      = datetime.fromtimestamp(ts / 1000)
day_str = dt.strftime(”%Y-%m-%d”)
for emp in employees:
if emp[“name_en”].lower() in username.lower():
d = result[emp[“name_en”]]
if day_str not in d:
d[day_str] = {“first”: dt, “last”: dt}
else:
if dt < d[day_str][“first”]: d[day_str][“first”] = dt
if dt > d[day_str][“last”]:  d[day_str][“last”]  = dt
return result

# ============================================================

# منطق المخالفات

# ============================================================

LATE_TABLE = [
(1,  15,  [0, .05, .10, .20], True,  “م1”),
(15, 30,  [.10,.15,.25,.50],  False, “م3”),
(30, 60,  [.25,.50,.75,1.0],  False, “م5”),
(60, 999, [1.0,2.0,3.0,3.0],  True,  “م7”),
]
EARLY_TABLE = [
(15,  [0,.10,.25,1.0],  True,  “م8”),
(999, [.10,.25,.50,1.0],False, “م9”),
]

def get_viol(minutes, occ, vtype=“late”):
idx   = min(occ, 4) - 1
table = LATE_TABLE if vtype == “late” else EARLY_TABLE
for entry in table:
if vtype == “late”:
mn, mx, rates, warn, code = entry
if mn <= minutes <= mx:
return code, rates[idx], warn and occ == 1
else:
mx, rates, warn, code = entry
if minutes <= mx:
return code, rates[idx], warn and occ == 1
return None, 0, False

def count_occ(emp_name, month_str, vtype):
with get_db() as db:
col = “delay_minutes” if vtype == “late” else “early_minutes”
row = db.execute(
f”SELECT COUNT(*) as n FROM attendance WHERE emp_name_en=? AND date LIKE ? AND {col}>0”,
(emp_name, f”{month_str}%”)
).fetchone()
return (row[“n”] or 0) + 1

# ============================================================

# المعالجة اليومية

# ============================================================

def process_day(target: date = None):
if target is None: target = date.today()
print(f”\n=== معالجة يوم {target} ===”)
with get_db() as db:
employees = [dict(r) for r in db.execute(“SELECT * FROM employees WHERE active=1”).fetchall()]
if not employees: return

```
try:
    records = fetch_ttlock_records(target, target)
except Exception as e:
    print(f"خطأ TTLock: {e}"); return

daily = parse_daily(records, employees)
day_str = target.strftime("%Y-%m-%d")
month_str = target.strftime("%Y-%m")

with get_db() as db:
    for emp in employees:
        en   = emp["name_en"]
        times = daily.get(en, {}).get(day_str)
        if not times:
            print(f"  لا بصمة: {emp['name_ar']}"); continue

        ci = times["first"]
        co = times["last"] if times["last"] != times["first"] else None

        if emp["type"] == "fixed":
            sh, sm = map(int, emp["shift_start"].split(":"))
            eh, em_h = map(int, emp["shift_end"].split(":"))
            exp_in  = ci.replace(hour=sh, minute=sm, second=0, microsecond=0)
            exp_out = co.replace(hour=eh, minute=em_h, second=0, microsecond=0) if co else None

            dm = round((ci - exp_in).total_seconds() / 60, 1)
            em = round((exp_out - co).total_seconds() / 60, 1) if exp_out and co else 0

            late_occ  = count_occ(en, month_str, "late")
            early_occ = count_occ(en, month_str, "early")

            vi, pi, wi = get_viol(dm, late_occ)  if dm > 0 else (None, 0, False)
            vo, po, wo = get_viol(em, early_occ, "early") if em > 0 else (None, 0, False)
            ded = round((pi + po) * emp["salary"], 2)

            db.execute("""
                INSERT OR REPLACE INTO attendance
                (emp_name_en,date,check_in,check_out,delay_minutes,early_minutes,
                 violation_in,violation_out,deduction_sar,warning)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (en, day_str,
                  ci.strftime("%H:%M:%S"),
                  co.strftime("%H:%M:%S") if co else None,
                  dm if dm > 0 else 0, em if em > 0 else 0,
                  vi, vo, ded, 1 if (wi or wo) else 0))

            send_notify(emp, target, ci, co, dm, em, vi, vo, ded)
            print(f"  {emp['name_ar']}: {'تأخر '+str(round(dm))+' د' if dm>0 else 'في الوقت'}")

        else:  # flex — إشعار نهاية الأسبوع
            if target.weekday() == 4:  # الجمعة
                ws_d = target - timedelta(days=(target.weekday()+2)%7)
                we_d = ws_d + timedelta(days=6)
                week_times = {
                    d: t for d, t in daily.get(en, {}).items()
                    if ws_d <= datetime.strptime(d,"%Y-%m-%d").date() <= we_d
                }
                actual = sum((t["last"]-t["first"]).total_seconds()/3600 for t in week_times.values())
                req    = emp.get("weekly_hours", 40)
                diff   = actual - req
                rate   = emp["salary"] / (req * 4)
                ded    = round(abs(diff)*rate, 2) if diff < 0 else 0
                send_notify_flex(emp, ws_d, we_d, actual, req, ded)
```

# ============================================================

# الإيميل

# ============================================================

def send_email(to, subject, html):
cfg = EMAIL_CONFIG
msg = MIMEMultipart(“alternative”)
msg[“Subject”] = subject
msg[“From”]    = cfg[“sender_email”]
msg[“To”]      = to
msg.attach(MIMEText(html, “html”, “utf-8”))
with smtplib.SMTP(cfg[“smtp_server”], cfg[“smtp_port”]) as s:
s.starttls()
s.login(cfg[“sender_email”], cfg[“sender_password”])
s.sendmail(cfg[“sender_email”], to, msg.as_string())

def send_notify(emp, day, ci, co, dm, em, vi, vo, ded):
name    = emp[“name_ar”]
co_str  = co.strftime(”%H:%M:%S”) if co else “—”
on_time = dm <= 0 and em <= 0
if on_time:
subj = “شكراً على الالتزام بالدوام ✅”
body = f”””<div dir="rtl" style="font-family:Arial;font-size:14px;color:#1F4E79">

<h2>أهلاً {name}،</h2>
<p>نشكرك على الالتزام بمواعيد الدوام يوم <b>{day}</b>.</p>
<ul><li>الحضور: <b>{ci.strftime("%H:%M:%S")}</b></li>
<li>الانصراف: <b>{co_str}</b></li></ul>
<p>نقدّر التزامك 🌟</p></div>"""
    else:
        viols = ""
        if dm > 0: viols += f"<li>تأخر <b>{round(dm)} دقيقة</b> — {vi or ''}</li>"
        if em > 0: viols += f"<li>مغادرة مبكرة <b>{round(em)} دقيقة</b> — {vo or ''}</li>"
        subj = "إشعار مخالفة دوام ⚠️"
        body = f"""<div dir="rtl" style="font-family:Arial;font-size:14px">
<h2 style="color:#C00000">إشعار مخالفة — {name}</h2>
<p>التاريخ: <b>{day}</b></p>
<ul><li>الحضور: <b>{ci.strftime("%H:%M:%S")}</b></li>
<li>الانصراف: <b>{co_str}</b></li></ul>
<h3 style="color:#C00000">المخالفات:</h3><ul>{viols}</ul>
{"<p><b>الخصم: "+str(ded)+" ريال</b></p>" if ded>0 else ""}
</div>"""
    try:
        send_email(emp["email"], subj, body)
    except Exception as e:
        print(f"  فشل إيميل {name}: {e}")

def send_notify_flex(emp, ws, we, actual, req, ded):
name = emp[“name_ar”]
diff = actual - req
if diff >= 0:
subj = “أتممت ساعاتك الأسبوعية ✅”
body = f”””<div dir="rtl" style="font-family:Arial;font-size:14px;color:#1F4E79">

<h2>أهلاً {name}،</h2>
<p>أتممت ساعاتك من <b>{ws}</b> إلى <b>{we}</b>.</p>
<ul><li>المنجز: <b>{round(actual,1)} ساعة</b></li>
<li>المطلوب: <b>{req} ساعة</b></li></ul>
<p>نقدّر جهدك 🌟</p></div>"""
    else:
        subj = "ساعاتك الأسبوعية غير مكتملة ⚠️"
        body = f"""<div dir="rtl" style="font-family:Arial;font-size:14px">
<h2 style="color:#C00000">إشعار — {name}</h2>
<p>الفترة: <b>{ws}</b> إلى <b>{we}</b></p>
<ul><li>المنجز: <b>{round(actual,1)} ساعة</b></li>
<li>المطلوب: <b>{req} ساعة</b></li>
<li>الناقص: <b>{round(abs(diff),1)} ساعة</b></li></ul>
<p><b>الخصم: {ded} ريال</b></p></div>"""
    try:
        send_email(emp["email"], subj, body)
    except Exception as e:
        print(f"  فشل إيميل {name}: {e}")

# ============================================================

# Excel Export

# ============================================================

def export_excel(month_str):
thin  = Side(border_style=“thin”, color=“BFBFBF”)
BRD   = Border(left=thin,right=thin,top=thin,bottom=thin)
HFILL = PatternFill(“solid”, start_color=“1F4E79”)
SFILL = PatternFill(“solid”, start_color=“2E75B6”)
LFILL = PatternFill(“solid”, start_color=“FFE0E0”)
GFILL = PatternFill(“solid”, start_color=“E2EFDA”)
WFONT = Font(name=“Arial”, bold=True, color=“FFFFFF”, size=10)
NFONT = Font(name=“Arial”, size=10)
BFONT = Font(name=“Arial”, bold=True, size=10)
CALIGN = Alignment(horizontal=“center”, vertical=“center”, wrap_text=True)

```
def cell(ws, r, col, val, font=None, fill=None):
    c = ws.cell(row=r, column=col, value=val)
    c.font = font or NFONT; c.fill = fill or PatternFill()
    c.alignment = CALIGN; c.border = BRD; return c

wb = openpyxl.Workbook()
if "Sheet" in wb.sheetnames: del wb["Sheet"]

with get_db() as db:
    employees = [dict(r) for r in db.execute("SELECT * FROM employees WHERE active=1").fetchall()]
    att_rows  = [dict(r) for r in db.execute(
        "SELECT * FROM attendance WHERE date LIKE ? ORDER BY emp_name_en,date",
        (f"{month_str}%",)).fetchall()]

months_ar = ["يناير","فبراير","مارس","أبريل","مايو","يونيو",
             "يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر"]
y, m = map(int, month_str.split("-"))
mlabel = f"{months_ar[m-1]} {y}"

# sheet لكل موظف
for emp in employees:
    en   = emp["name_en"]
    rows = [r for r in att_rows if r["emp_name_en"] == en]
    ws   = wb.create_sheet(en)
    ws.sheet_view.rightToLeft = True

    ws.merge_cells("A1:J1")
    ws["A1"].value = f"الحضور والغياب — {emp['name_ar']} — {mlabel}"
    ws["A1"].font  = Font(name="Arial", bold=True, size=13, color="1F4E79")
    ws["A1"].alignment = CALIGN

    hdrs = ["التاريخ","الحضور","الانصراف","تأخر(د)","مبكر(د)","مخالفة حضور","مخالفة انصراف","إنذار","الخصم(ر)","الراتب"]
    for c_i, h in enumerate(hdrs, 1):
        cell(ws, 2, c_i, h, WFONT, SFILL)

    for i, r in enumerate(rows, 3):
        fill = LFILL if r["delay_minutes"] > 0 or r["early_minutes"] > 0 else GFILL
        cell(ws, i, 1,  r["date"])
        cell(ws, i, 2,  r["check_in"] or "—",   fill=fill)
        cell(ws, i, 3,  r["check_out"] or "—",  fill=fill)
        cell(ws, i, 4,  r["delay_minutes"] or 0, fill=fill)
        cell(ws, i, 5,  r["early_minutes"] or 0, fill=fill)
        cell(ws, i, 6,  r["violation_in"] or "", fill=fill)
        cell(ws, i, 7,  r["violation_out"] or "",fill=fill)
        cell(ws, i, 8,  "نعم" if r["warning"] else "", fill=fill)
        cell(ws, i, 9,  r["deduction_sar"] or 0, fill=fill)
        cell(ws, i, 10, emp["salary"],             fill=fill)

    last = len(rows) + 3
    ws.merge_cells(f"A{last}:H{last}")
    cell(ws, last, 1, "إجمالي الخصومات الشهرية", BFONT, SFILL)
    cell(ws, last, 9, f"=SUM(I3:I{last-1})", BFONT, SFILL)

    for i, w in enumerate([13,10,10,10,10,14,14,8,12,12],1):
        ws.column_dimensions[get_column_letter(i)].width = w

# sheet مسيرة الرواتب
ws_pay = wb.create_sheet("مسيرة الرواتب", 0)
ws_pay.sheet_view.rightToLeft = True
ws_pay.merge_cells("A1:L1")
ws_pay["A1"].value = f"مسيرة الرواتب — {mlabel}"
ws_pay["A1"].font  = Font(name="Arial", bold=True, size=15, color="FFFFFF")
ws_pay["A1"].fill  = HFILL
ws_pay["A1"].alignment = CALIGN
ws_pay.row_dimensions[1].height = 34

pay_hdrs = ["م","الموظف","الأساسي","بدل سكن","بدل مواصلات","إجمالي البدلات",
            "عمولة","إجمالي المستحقات","خصم حضور","خصومات أخرى","إجمالي الخصم","صافي الراتب"]
for c_i, h in enumerate(pay_hdrs, 1):
    cell(ws_pay, 2, c_i, h, WFONT, SFILL)
ws_pay.row_dimensions[2].height = 26

for idx, emp in enumerate(employees, 1):
    r      = idx + 2
    en     = emp["name_en"]
    ded    = sum(a["deduction_sar"] for a in att_rows if a["emp_name_en"] == en)
    fill   = PatternFill("solid", start_color="EBF3FB") if idx%2==0 else PatternFill("solid",start_color="FFFFFF")
    values = [idx, emp["name_ar"], emp["salary"],
              emp["housing_allowance"], emp["transport_allowance"],
              f"=D{r}+E{r}", emp["commission"],
              f"=C{r}+F{r}+G{r}", round(ded,2), emp["other_deductions"],
              f"=I{r}+J{r}", f"=H{r}-K{r}"]
    for c_i, v in enumerate(values, 1):
        cl = cell(ws_pay, r, c_i, v, fill=fill)
        if c_i == 12: cl.font = Font(name="Arial", size=10, bold=True, color="1F4E79")

last_r = len(employees) + 3
ws_pay.merge_cells(f"A{last_r}:B{last_r}")
cell(ws_pay, last_r, 1, "الإجمالي", BFONT, PatternFill("solid",start_color="D6E4F0"))
for c_i in [3,4,5,9,10]:
    val = sum(emp[["salary","housing_allowance","transport_allowance","other_deductions","other_deductions"][c_i-3]] for emp in employees)
for c_i in [6,7,8,11,12]:
    cell(ws_pay, last_r, c_i, f"=SUM({get_column_letter(c_i)}3:{get_column_letter(c_i)}{last_r-1})",
         BFONT, PatternFill("solid",start_color="D6E4F0"))

for i, w in enumerate([4,16,12,11,13,13,11,15,12,13,13,13],1):
    ws_pay.column_dimensions[get_column_letter(i)].width = w

buf = io.BytesIO()
wb.save(buf)
buf.seek(0)
return buf
```

# ============================================================

# API Routes

# ============================================================

@app.route(”/”)
def index():
return render_template(“index.html”)

@app.route(”/api/dashboard”)
def api_dashboard():
today = date.today().strftime(”%Y-%m-%d”)
month = date.today().strftime(”%Y-%m”)
with get_db() as db:
emp_count = db.execute(“SELECT COUNT(*) as n FROM employees WHERE active=1”).fetchone()[“n”]
today_att = db.execute(
“SELECT COUNT(*) as n FROM attendance WHERE date=?”, (today,)).fetchone()[“n”]
today_late = db.execute(
“SELECT COUNT(*) as n FROM attendance WHERE date=? AND delay_minutes>0”, (today,)).fetchone()[“n”]
month_ded  = db.execute(
“SELECT COALESCE(SUM(deduction_sar),0) as s FROM attendance WHERE date LIKE ?”,
(f”{month}%”,)).fetchone()[“s”]
recent = [dict(r) for r in db.execute(”””
SELECT a.*, e.name_ar FROM attendance a
JOIN employees e ON a.emp_name_en = e.name_en
ORDER BY a.date DESC, a.created_at DESC LIMIT 10
“””).fetchall()]
return jsonify({
“emp_count”: emp_count, “today_att”: today_att,
“today_late”: today_late, “month_deductions”: round(month_ded, 2),
“recent”: recent
})

@app.route(”/api/employees”, methods=[“GET”])
def api_get_employees():
with get_db() as db:
rows = [dict(r) for r in db.execute(“SELECT * FROM employees WHERE active=1 ORDER BY name_ar”).fetchall()]
return jsonify(rows)

@app.route(”/api/employees”, methods=[“POST”])
def api_add_employee():
d = request.json
with get_db() as db:
db.execute(”””
INSERT INTO employees (name_en,name_ar,email,salary,type,shift_start,shift_end,
weekly_hours,housing_allowance,transport_allowance,commission,other_deductions,lock_id)
VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
“””, (d[“name_en”], d[“name_ar”], d.get(“email”,””), d.get(“salary”,0),
d.get(“type”,“fixed”), d.get(“shift_start”,“09:00”), d.get(“shift_end”,“17:00”),
d.get(“weekly_hours”,40), d.get(“housing_allowance”,0),
d.get(“transport_allowance”,0), d.get(“commission”,0),
d.get(“other_deductions”,0), d.get(“lock_id”,””)))
return jsonify({“ok”: True})

@app.route(”/api/employees/<int:emp_id>”, methods=[“PUT”])
def api_update_employee(emp_id):
d = request.json
with get_db() as db:
db.execute(”””
UPDATE employees SET name_ar=?,email=?,salary=?,type=?,shift_start=?,shift_end=?,
weekly_hours=?,housing_allowance=?,transport_allowance=?,commission=?,other_deductions=?
WHERE id=?
“””, (d[“name_ar”], d[“email”], d[“salary”], d[“type”],
d[“shift_start”], d[“shift_end”], d[“weekly_hours”],
d[“housing_allowance”], d[“transport_allowance”],
d[“commission”], d[“other_deductions”], emp_id))
return jsonify({“ok”: True})

@app.route(”/api/employees/<int:emp_id>”, methods=[“DELETE”])
def api_delete_employee(emp_id):
with get_db() as db:
db.execute(“UPDATE employees SET active=0 WHERE id=?”, (emp_id,))
return jsonify({“ok”: True})

@app.route(”/api/attendance”)
def api_attendance():
month = request.args.get(“month”, date.today().strftime(”%Y-%m”))
emp   = request.args.get(“emp”, “”)
query = “SELECT a.*, e.name_ar FROM attendance a JOIN employees e ON a.emp_name_en=e.name_en WHERE a.date LIKE ?”
params = [f”{month}%”]
if emp:
query += “ AND a.emp_name_en=?”; params.append(emp)
query += “ ORDER BY a.date DESC”
with get_db() as db:
rows = [dict(r) for r in db.execute(query, params).fetchall()]
return jsonify(rows)

@app.route(”/api/run_today”, methods=[“POST”])
def api_run_today():
try:
process_day()
return jsonify({“ok”: True, “msg”: “تمت المعالجة بنجاح”})
except Exception as e:
return jsonify({“ok”: False, “msg”: str(e)}), 500

@app.route(”/api/export_excel”)
def api_export_excel():
month = request.args.get(“month”, date.today().strftime(”%Y-%m”))
buf   = export_excel(month)
months_ar = [“يناير”,“فبراير”,“مارس”,“أبريل”,“مايو”,“يونيو”,
“يوليو”,“أغسطس”,“سبتمبر”,“أكتوبر”,“نوفمبر”,“ديسمبر”]
y, m  = map(int, month.split(”-”))
fname = f”attendance_{months_ar[m-1]}_{y}.xlsx”
return send_file(buf, as_attachment=True, download_name=fname,
mimetype=“application/vnd.openxmlformats-officedocument.spreadsheetml.sheet”)

# ============================================================

# Scheduler — يشتغل تلقائياً كل يوم الساعة 8 مساءً

# ============================================================

scheduler = BackgroundScheduler()
scheduler.add_job(process_day, “cron”, hour=20, minute=0)
scheduler.start()

if **name** == “**main**”:
init_db()
app.run(host=“0.0.0.0”, port=int(os.environ.get(“PORT”, 5000)), debug=False)
